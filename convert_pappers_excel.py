import os
import requests
import pandas as pd
from fastapi import HTTPException
from openpyxl import load_workbook

# --- On force Python à travailler dans le dossier du script ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(SCRIPT_DIR)
print("Dossier courant :", os.getcwd())

from config import settings

# Lire la clé et l'URL depuis la configuration (variables d'environnement)
PAPPERS_API_KEY = settings.PAPPERS_API_KEY
PAPPERS_URL = settings.PAPPERS_BASE_URL or "https://api.pappers.fr/v2/entreprise"
PAPPERS_TIMEOUT = getattr(settings, "PAPPERS_TIMEOUT_SECONDS", 5)


# ---------------------------------------------------------
# API Pappers
# ---------------------------------------------------------
def fetch_pappers_data(siret: str) -> dict:
    if len(siret) != 14 or not siret.isdigit():
        raise HTTPException(400, "Le SIRET doit contenir exactement 14 chiffres")

    params = {"api_token": PAPPERS_API_KEY, "siret": siret}

    if not PAPPERS_API_KEY:
        raise HTTPException(500, "Clé Pappers non configurée (PAPPERS_API_KEY)")

    try:
        response = requests.get(PAPPERS_URL, params=params, timeout=PAPPERS_TIMEOUT)
        response.raise_for_status()
    except requests.exceptions.Timeout:
        raise HTTPException(504, "Timeout lors de l'appel à l'API Pappers")
    except requests.exceptions.HTTPError:
        status = getattr(response, "status_code", None)
        detail = getattr(response, "text", "")
        raise HTTPException(502, f"Erreur Pappers HTTP {status}: {detail}")
    except requests.exceptions.RequestException as e:
        raise HTTPException(502, f"Erreur réseau vers Pappers: {e}")

    return response.json()


# ---------------------------------------------------------
# Sauvegarde TXT
# ---------------------------------------------------------
def save_to_txt(siret: str, data: dict):
    filename = f"{siret}.txt"
    with open(filename, "w", encoding="utf-8") as f:
        f.write(str(data))
    print(f"✔ Fichier généré : {filename}")


# ---------------------------------------------------------
# Conversion TXT → DataFrame
# ---------------------------------------------------------
def parse_pappers_text(path):
    if not os.path.exists(path):
        print(f"❌ Fichier introuvable : {path}")
        return None

    with open(path, "r", encoding="utf-8") as f:
        content = f.read()

    start = content.find("{'siren': '")
    if start != -1:
        content = content[start:]

    content = content.replace("\n", " ")
    items = content.split(",")

    rows = []
    for item in items:
        if ":" not in item:
            continue
        key, value = item.split(":", 1)
        rows.append([key.strip().replace("'", ""), value.strip().replace("'", "")])

    df = pd.DataFrame(rows, columns=["Champ", "Valeur"])

    if df.empty:
        df = pd.DataFrame([["Aucune donnée", ""]], columns=["Champ", "Valeur"])

    df.insert(0, "Etab", "")
    df.insert(1, "Siege", "")

    return df


# ---------------------------------------------------------
# Programme principal
# ---------------------------------------------------------
if __name__ == "__main__":
    print("=== Recherche interactive multi-SIRET Pappers ===\n")

    siret_memory = []

    # --- Saisie interactive ---
    while True:
        siret = input("Entrez un numéro de SIRET (ou Entrée pour terminer la saisie) : ").strip()

        if not siret:
            break

        siret_memory.append(siret)
        print(f"✔ SIRET ajouté : {siret}")

        again = input("Voulez-vous entrer un autre SIRET ? (o/n) : ").strip().lower()
        if again != "o":
            break

    if not siret_memory:
        print("❌ Aucun SIRET fourni → aucun fichier généré.")
        exit()

    # --- Génération TXT ---
    print("\n=== Génération des fichiers .txt ===")
    for siret in siret_memory:
        try:
            data = fetch_pappers_data(siret)
            save_to_txt(siret, data)
        except Exception as e:
            print(f"❌ Erreur pour {siret} : {e}")

    # --- Génération Excel ---
    print("\n=== Génération du classeur Excel ===")

    base_name = "pappers_classeur"
    extension = ".xlsx"
    output_file = base_name + extension

    i = 1
    while os.path.exists(output_file):
        output_file = f"{base_name}_{i}{extension}"
        i += 1

    print("Nom du fichier généré :", output_file)

    # Création du classeur
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for siret in siret_memory:
            txt_file = f"{siret}.txt"
            df = parse_pappers_text(txt_file)

            if df is None:
                continue

            sheet_name = siret[-10:]  # Excel safe
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Mise en forme
    wb = load_workbook(output_file)

    for siret in siret_memory:
        sheet_name = siret[-10:]
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        ws.auto_filter.ref = "A1:E1"
        ws.freeze_panes = "A2"
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 50

    wb.save(output_file)

    print("\n✔ Tous les fichiers TXT + le classeur Excel multi-onglets ont été générés.")
